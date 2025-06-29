
"""streamlit_app.py (v3.0 – refonte complète)
================================================
Application Streamlit pour gérer les tirages au sort hebdomadaires à partir
du fichier **Liste_membres_Train.xlsx**.

Fonctionnalités clés
--------------------
1. **Source unique** : le classeur Excel contient la feuille « Membres » (données
   d’origine) *et* la feuille « Tirages » (historique). Le code lit / écrit dans
   ce même fichier → persistance garantie.
2. **Exclusions automatiques** : si la colonne **Motif sortie** d’un joueur est
   renseignée, il est ignoré du tirage.
3. **Date du train** : lorsqu’un joueur est désigné *Titulaire*, la date
   correspondante est ajoutée/complétée dans sa colonne **Date du train** (une
   liste de dates séparées par virgule si plusieurs tirages).
4. **Tirage hebdomadaire libre** : choisis une semaine future (lundi…) qui n’a
   pas encore été tirée. Titulaires + Suppléants sont générés ➜ enregistrés dans
   la feuille « Tirages » + colonne Date du train.
5. **Modification manuelle** : ouvre un planning dans un formulaire, change un
   Titulaire (ou Suppléant) puis *Enregistrer* → met à jour historique + Excel.
6. **Historique complet** : toutes les semaines déjà tirées sont listées avec
   leurs tableaux.
7. **Reset sécurisé** : bouton « Réinitialiser tirages » dans la barre latérale
   ➜ demande de taper « CONFIRMER » avant d’effacer la feuille « Tirages » et
   de vider toutes les dates du train.

Dépendances (requirements.txt)
------------------------------
```
streamlit>=1.35
pandas
openpyxl
```
"""
from __future__ import annotations

import datetime as dt
import random
import sqlite3  # used only for in-memory manipulation; no persistent DB now
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl
import pandas as pd
import streamlit as st

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

DATA_FILE = Path("Liste_membres_Train.xlsx")
MEMBRES_SHEET = "Membres"
TIRAGES_SHEET = "Tirages"
WEEKS_AHEAD_SHOWN = 52  # nombre de semaines futures affichées dans la liste

# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def load_workbook() -> openpyxl.Workbook:
    if not DATA_FILE.exists():
        st.error(f"Fichier {DATA_FILE} introuvable. Téléverse-le puis relance l'app.")
        st.stop()
    return openpyxl.load_workbook(DATA_FILE)


def load_players_df(wb: openpyxl.Workbook) -> pd.DataFrame:
    if MEMBRES_SHEET not in wb.sheetnames:
        st.error(f"La feuille '{MEMBRES_SHEET}' est manquante dans {DATA_FILE}.")
        st.stop()
    df = pd.read_excel(DATA_FILE, sheet_name=MEMBRES_SHEET, engine="openpyxl")
    # Nettoyage des colonnes attendues
    expected_cols = {"Pseudo", "Motif sortie", "Date du train"}
    if not expected_cols.issubset(df.columns):
        st.error("Le fichier doit contenir les colonnes : " + ", ".join(expected_cols))
        st.stop()
    return df


def save_players_df(df: pd.DataFrame, wb: openpyxl.Workbook) -> None:
    with pd.ExcelWriter(DATA_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, index=False, sheet_name=MEMBRES_SHEET)


def append_tirages_rows(rows: List[Tuple[str, str, str]], wb: openpyxl.Workbook) -> None:
    """Append rows to sheet (create if absent).
    rows = [(week_id, date_iso, titulaire_pseudo, suppleant_pseudo)]
    """
    if TIRAGES_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(TIRAGES_SHEET)
        ws.append(["Semaine", "Date", "Titulaire", "Suppléant"])
    else:
        ws = wb[TIRAGES_SHEET]
    for r in rows:
        ws.append(list(r))
    wb.save(DATA_FILE)


def clear_tirages_sheet_and_dates(wb: openpyxl.Workbook) -> None:
    # Delete Tirages sheet if exists
    if TIRAGES_SHEET in wb.sheetnames:
        wb.remove(wb[TIRAGES_SHEET])
    # Clear Date du train column
    df = load_players_df(wb)
    if "Date du train" in df.columns:
        df["Date du train"] = pd.NA
    save_players_df(df, wb)
    wb.save(DATA_FILE)

# ---------------------------------------------------------------------------
# Utility helpers for weeks & ids
# ---------------------------------------------------------------------------

def week_id_for_date(d: dt.date) -> str:
    year, week, _ = d.isocalendar()
    return f"{year}-W{week:02d}"


def monday_of_week(d: dt.date) -> dt.date:
    return d - dt.timedelta(days=d.weekday())


def upcoming_week_mondays(n: int = WEEKS_AHEAD_SHOWN) -> List[dt.date]:
    today = dt.date.today()
    next_monday = monday_of_week(today + dt.timedelta(days=7))
    return [next_monday + dt.timedelta(days=7 * i) for i in range(n)]

# ---------------------------------------------------------------------------
# Tirage engine (in-memory only, persistence via Excel sheet)
# ---------------------------------------------------------------------------

def eligible_players(df: pd.DataFrame) -> pd.DataFrame:
    return df[df["Motif sortie"].fillna("").str.strip() == ""]


def draw_week(df_players: pd.DataFrame, week_monday: dt.date) -> Dict[dt.date, Tuple[str, str]]:
    dates = [week_monday + dt.timedelta(i) for i in range(7)]
    pseudos = df_players["Pseudo"].tolist()
    random.shuffle(pseudos)
    schedule: Dict[dt.date, Tuple[str, str]] = {}
    used_titulars: set[str] = set()
    pseudo_iter = iter(pseudos)
    for d in dates:
        tit = next(p for p in pseudo_iter if p not in used_titulars)
        used_titulars.add(tit)
        sup = next(p for p in pseudo_iter if p != tit)
        schedule[d] = (tit, sup)
    return schedule

# ---------------------------------------------------------------------------
# Data access for Tirages sheet
# ---------------------------------------------------------------------------

def tirages_df(wb: openpyxl.Workbook) -> pd.DataFrame:
    if TIRAGES_SHEET not in wb.sheetnames:
        return pd.DataFrame(columns=["Semaine", "Date", "Titulaire", "Suppléant"])
    return pd.read_excel(DATA_FILE, sheet_name=TIRAGES_SHEET, engine="openpyxl")


def week_exists(wb: openpyxl.Workbook, week_id: str) -> bool:
    df = tirages_df(wb)
    return week_id in df["Semaine"].unique()

# ---------------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------------

st.set_page_config(page_title="Tirage au sort train", page_icon="🎲", layout="centered")

st.title("🎲 Tirages au sort – Liste Train")

wb = load_workbook()
players_df = load_players_df(wb)

# ---- Sidebar : Générer une nouvelle semaine ---------------------------------------------------

st.sidebar.header("Générer une nouvelle semaine")

existing_week_ids = set(tirages_df(wb)["Semaine"].unique())
week_options = [m for m in upcoming_week_mondays() if week_id_for_date(m) not in existing_week_ids]

if week_options:
    monday_sel = st.sidebar.selectbox(
        "Choisis la semaine (lundi)",
        week_options,
        format_func=lambda d: f"Semaine {week_id_for_date(d)} (débute le {d.strftime('%d/%m/%Y')})",
    )
    if st.sidebar.button("🎲 Générer cette semaine"):
        elig_df = eligible_players(players_df)
        if len(elig_df) < 14:  # 7 jours × 2 joueurs minimum
            st.sidebar.error("Pas assez de joueurs éligibles pour générer une semaine complète !")
        else:
            schedule = draw_week(elig_df, monday_sel)

            # 1. Historique (feuille Tirages)
            rows = [
                (week_id_for_date(monday_sel), d.isoformat(), tit, sup)
                for d, (tit, sup) in schedule.items()
            ]
            append_tirages_rows(rows, wb)

            # 2. Date du train (ajout / concat)
            date_map = {tit: d.isoformat() for d, (tit, _) in schedule.items()}
            players_df["Date du train"] = players_df.apply(
                lambda r: _concat_date(r["Date du train"], date_map.get(r["Pseudo"])) if r["Pseudo"] in date_map else r["Date du train"],
                axis=1,
            )
            save_players_df(players_df, wb)

            st.sidebar.success("✅ Semaine générée et enregistrée !")
            st.experimental_rerun()
else:
    st.sidebar.info("Toutes les semaines futures déjà générées (jusqu'à 1 an).")

# ---- Sidebar : Réinitialisation ---------------------------------------------------------------

st.sidebar.header("Réinitialiser les tirages")
if st.sidebar.button("🗑️ Réinitialiser (tests)"):
    confirm = st.sidebar.text_input("Écris CONFIRMER pour valider")
    if confirm == "CONFIRMER":
        clear_tirages_sheet_and_dates(wb)
        st.sidebar.success("Tous les tirages ont été réinitialisés !")
        st.experimental_rerun()
    else:
        st.sidebar.warning("Action annulée – saisie incorrecte.")

# ---- Fonction pour concaténer dates dans la colonne Date du train -----------------------------

def _concat_date(existing: str | float | None, new_date: str | None) -> str | None:
    if new_date is None:
        return existing
    if pd.isna(existing) or existing is None:
        return new_date
    existing_str = str(existing).strip()
    if new_date in existing_str.split(","):
        return existing_str  # déjà présent
    return existing_str + ", " + new_date

# ---- Affichage de l'historique ---------------------------------------------------------------

st.subheader("Historique des semaines tirées")

tir_df = tirages_df(wb)
if tir_df.empty:
    st.info("Aucun tirage enregistré pour l'instant.")
else:
    for wid in sorted(tir_df["Semaine"].unique()):
        with st.expander(f"Semaine {wid}"):
            week_df = tir_df[tir_df["Semaine"] == wid][["Date", "Titulaire", "Suppléant"]].copy()
            week_df["Date"] = pd.to_datetime(week_df["Date"]).dt.strftime("%A %d/%m/%Y")
            week_df.set_index("Date", inplace=True)

            # Editable Titulaire / Suppléant
            edited = st.experimental_data_editor(week_df, key=f"edit_{wid}")
            if st.button("💾 Enregistrer les modifications", key=f"save_{wid}"):
                _apply_edits_and_save(edited, wid, wb, players_df)
                st.success("Modifications enregistrées !")
                st.experimental_rerun()

# ---------------------------------------------------------------------------
# Helper to apply manual edits
# ---------------------------------------------------------------------------

def _apply_edits_and_save(edited_df: pd.DataFrame, week_id: str, wb: openpyxl.Workbook, players_df: pd.DataFrame) -> None:
    """Update Tirages sheet + Date du train column according to edited DataFrame."""
    # 1. Replace rows in Tirages sheet for that week
    if TIRAGES_SHEET not in wb.sheetnames:
        st.error("Feuille Tirages manquante – impossible de mettre à jour.")
        return
    ws = wb[TIRAGES_SHEET]
    # Remove existing rows of the week
    rows_to_delete = [idx for idx, row in enumerate(ws.iter_rows(values_only=True), start=1) if idx > 1 and row[0] == week_id]
    for excel_idx in reversed(rows_to_delete):
        ws.delete_rows(excel_idx)
    # Append new rows
    for date_str, row in edited_df.iterrows():
        iso_date = dt.datetime.strptime(date_str, "%A %d/%m/%Y").date().isoformat()
        ws.append([week_id, iso_date, row["Titulaire"], row["Suppléant"]])
    wb.save(DATA_FILE)

    # 2. Rebuild Date du train column : clear previous dates of this week then append new ones
    monday = dt.datetime.strptime(week_id + "-1", "%Y-W%W-%w").date()  # Monday date reconstruction
    week_dates = {monday + dt.timedelta(i) for i in range(7)}
    players_df["Date du train"] = players_df["Date du train"].apply(
        lambda x: _remove_week_dates(str(x), week_dates)
    )
    # Add new dates for titulaires only
    date_map = {row["Titulaire"]: iso for iso, row in zip(edited_df.index, edited_df.itertuples())}
    players_df["Date du train"] = players_df.apply(
        lambda r: _concat_date(r["Date du train"], date_map.get(r["Pseudo"])) if r["Pseudo"] in date_map else r["Date du train"],
        axis=1,
    )
    save_players_df(players_df, wb)


def _remove_week_dates(existing: str | None, week_dates: set[dt.date]) -> str | None:
    if pd.isna(existing) or existing is None:
        return existing
    dates = [dt.date.fromisoformat(d.strip()) for d in str(existing).split(",")]
    dates = [d for d in dates if d not in week_dates]
    return ", ".join(d.isoformat() for d in dates) if dates else None
